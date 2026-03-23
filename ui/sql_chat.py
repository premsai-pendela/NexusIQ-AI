"""
NexusIQ AI — SQL Agent Chat Interface
"""

import streamlit as st
import pandas as pd
import time
import random
import sys
from datetime import datetime
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent))

from agents.sql_agent import SQLAgent
from utils.validators import VALID_REGIONS, VALID_CATEGORIES


def run_sql_chat():
    """Main function for SQL Chat interface"""
    
    # ═══════════════════════════════════════════════════════
    #  CONFIGURATION
    # ═══════════════════════════════════════════════════════
    
    INSIGHTS = [
        "💡 **Did You Know?** We analyze 100,000 transactions across 5 regions",
        "🧠 **Did You Know?** Complex queries use Gemini 2.5 Pro — our smartest model",
        "⚡ **Did You Know?** Simple queries complete in under 5 seconds",
        "🔄 **Did You Know?** We auto-switch models if one hits quota limits",
        "🔒 **Did You Know?** All queries are read-only — your data stays safe",
        "📊 **Did You Know?** You can download results as CSV with one click",
        "🎯 **Did You Know?** Adding time ranges makes queries more precise",
        "🚀 **Did You Know?** Our circuit breaker skips failed models instantly",
        "💰 **Did You Know?** The database tracks $139M+ in total revenue",
        "🌐 **Did You Know?** We support 5 regions: East, West, North, South, Central",
        "🛒 **Did You Know?** Products span 5 categories: Electronics, Clothing, Food, Home, Sports",
        "⏱️ **Did You Know?** First query may be slower due to model warm-up",
    ]
    
    # ═══════════════════════════════════════════════════════
    #  HELPER FUNCTIONS
    # ═══════════════════════════════════════════════════════
    
    def format_time(seconds: float) -> str:
        if seconds < 60:
            return f"{int(seconds)} seconds"
        elif seconds < 3600:
            mins = int(seconds // 60)
            secs = int(seconds % 60)
            if secs == 0:
                return f"{mins} minute{'s' if mins > 1 else ''}"
            return f"{mins} minute{'s' if mins > 1 else ''} {secs} seconds"
        else:
            hours = int(seconds // 3600)
            remaining = seconds % 3600
            mins = int(remaining // 60)
            secs = int(remaining % 60)
            result = f"{hours} hour{'s' if hours > 1 else ''}"
            if mins > 0:
                result += f" {mins} minute{'s' if mins > 1 else ''}"
            if secs > 0:
                result += f" {secs} seconds"
            return result
    
    def time_ago(timestamp: datetime) -> str:
        diff = datetime.now() - timestamp
        seconds = diff.total_seconds()
        if seconds < 60:
            return "just now"
        elif seconds < 3600:
            return f"{int(seconds // 60)}m ago"
        elif seconds < 86400:
            return f"{int(seconds // 3600)}h ago"
        else:
            return f"{int(seconds // 86400)}d ago"
    
    # ═══════════════════════════════════════════════════════
    #  INITIALIZE
    # ═══════════════════════════════════════════════════════
    
    @st.cache_resource
    def get_agent():
        return SQLAgent(mode="development")
    
    agent = get_agent()
    
    # ═══════════════════════════════════════════════════════
    #  SESSION STATE
    # ═══════════════════════════════════════════════════════
    
    if "query_history" not in st.session_state:
        st.session_state.query_history = []
    if "selected_history" not in st.session_state:
        st.session_state.selected_history = None
    if "next_question" not in st.session_state:
        st.session_state.next_question = ""
    if "input_key" not in st.session_state:
        st.session_state.input_key = 0
    if "auto_run" not in st.session_state:
        st.session_state.auto_run = False
    if "last_result" not in st.session_state:
        st.session_state.last_result = None
    if "last_question" not in st.session_state:
        st.session_state.last_question = None
    if "last_time" not in st.session_state:
        st.session_state.last_time = None
    
    def add_to_history(question, result, execution_time):
        st.session_state.query_history.insert(0, {
            "question": question,
            "query": result.get("query", ""),
            "answer": result.get("answer", ""),
            "explanation": result.get("explanation", ""),
            "results": result.get("results", []),
            "row_count": result.get("row_count", 0),
            "success": result.get("success", False),
            "error": result.get("error", ""),
            "model_used": result.get("model_used", ""),
            "time": execution_time,
            "timestamp": datetime.now()
        })
        st.session_state.query_history = st.session_state.query_history[:10]
    
    # ═══════════════════════════════════════════════════════
    #  PAGE LAYOUT
    # ═══════════════════════════════════════════════════════
    
    st.title("🗄️ SQL Agent — Database Query Interface")
    st.markdown("*Ask questions about your sales data in plain English*")
    
    # ═══════════════════════════════════════════════════════
    #  SIDEBAR
    # ═══════════════════════════════════════════════════════
    
    with st.sidebar:
        st.header("📊 Database Schema")
        
        with st.expander("📋 sales_transactions"):
            st.code("• transaction_date\n• region (5 regions)\n• store_id\n• product_category\n• product_name\n• quantity, unit_price\n• total_amount\n• customer_id\n• payment_method")
        
        with st.expander("👥 customers"):
            st.code("• customer_id\n• name, email, region\n• signup_date\n• total_purchases")
        
        st.markdown("---")
        st.subheader("💡 Examples")
        st.info("**Simple:**\n- Total revenue?\n- Sales in West region?\n\n**Complex:**\n- Top 5 products?\n- Compare regions?")
        
        st.markdown("---")
        st.subheader("📊 Model Status")
        quota_status = agent.get_quota_status()
        if quota_status:
            for model, status in quota_status.items():
                st.caption(f"{status['status']} {model.split('-')[0]}")
        else:
            st.caption("🟢 All models available")
        
        if st.button("🔄 Reset Quota Tracking"):
            agent.reset_quota_tracking()
            st.rerun()
        
        st.markdown("---")
        st.subheader("📜 Query History")
        
        if st.session_state.query_history:
            for i, item in enumerate(st.session_state.query_history):
                icon = "✅" if item["success"] else "❌"
                short = item["question"][:30] + "..." if len(item["question"]) > 30 else item["question"]
                if st.button(f"{icon} {short}", key=f"history_{i}", use_container_width=True):
                    st.session_state.selected_history = item
                    st.rerun()
                st.caption(f"⏱️ {item['time']:.1f}s • {time_ago(item['timestamp'])}")
            
            if st.button("🗑️ Clear History", use_container_width=True):
                st.session_state.query_history = []
                st.rerun()
        else:
            st.caption("No queries yet")
    
    # ═══════════════════════════════════════════════════════
    #  TEXT INPUT WITH DYNAMIC KEY (THE FIX!)
    # ═══════════════════════════════════════════════════════
    
    # Check if we have a corrected/prefilled question
    prefill = ""
    if st.session_state.next_question:
        prefill = st.session_state.next_question
        st.session_state.next_question = ""
        st.session_state.auto_run = True
    
    # DYNAMIC KEY forces Streamlit to create a NEW widget
    # New widget = value= parameter is respected!
    question = st.text_input(
        "💬 Ask a question:",
        value=prefill,
        placeholder="e.g., What was total revenue last month?",
        key=f"q_input_{st.session_state.input_key}"
    )
    
    col1, col2 = st.columns([1, 5])
    with col1:
        ask_button = st.button("🔍 Ask", type="primary", use_container_width=True)
    
    # Determine if we should run
    should_run = (ask_button and question) or (st.session_state.auto_run and question)
    if st.session_state.auto_run:
        st.session_state.auto_run = False
    
    # ═══════════════════════════════════════════════════════
    #  DISPLAY SELECTED HISTORY
    # ═══════════════════════════════════════════════════════
    
    if st.session_state.selected_history:
        item = st.session_state.selected_history
        st.info(f"📜 **Showing saved result for:** {item['question']}")
        
        if item["success"]:
            st.success(f"✅ Completed in **{format_time(item['time'])}** (cached)")
            st.markdown("---")
            st.markdown("### 💬 Answer")
            st.markdown(item.get("answer", "No answer available"))
            
            if item.get("query"):
                with st.expander("🔍 View SQL Query"):
                    st.code(item["query"], language="sql")
            if item.get("explanation"):
                with st.expander("📖 How This Query Works"):
                    st.markdown(item["explanation"])
            if item.get("results"):
                with st.expander(f"📊 View Data ({item.get('row_count', 0)} rows)"):
                    df = pd.DataFrame(item["results"])
                    st.dataframe(df, use_container_width=True)
                    st.markdown("**📥 Export:**")
                    h_col1, h_col2, h_col3 = st.columns(3)
                    with h_col1:
                        st.download_button("📄 CSV", data=df.to_csv(index=False), file_name="history_export.csv", mime="text/csv", use_container_width=True, key=f"h_csv_{time.time()}")
                    with h_col2:
                        st.download_button("📋 JSON", data=df.to_json(orient="records", indent=2), file_name="history_export.json", mime="application/json", use_container_width=True, key=f"h_json_{time.time()}")
                    with h_col3:
                        import io
                        buf = io.BytesIO()
                        df.to_excel(buf, index=False, engine='openpyxl')
                        st.download_button("📊 Excel", data=buf.getvalue(), file_name="history_export.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key=f"h_excel_{time.time()}")
        else:
            st.error(f"❌ This query failed: {item.get('error', '')}")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🔄 Re-run This Query"):
                st.session_state.next_question = item["question"]
                st.session_state.input_key += 1
                st.session_state.selected_history = None
                st.rerun()
        with col2:
            if st.button("❌ Close & Ask New"):
                st.session_state.selected_history = None
                st.rerun()
        
        st.stop()
    
    # ═══════════════════════════════════════════════════════
    #  QUERY PROCESSING
    # ═══════════════════════════════════════════════════════
    
    if should_run:
        st.session_state.selected_history = None
        start_time = time.time()
        result = None
        
        # Progress UI
        progress_container = st.container()
        with progress_container:
            progress_bar = st.progress(0)
            status_text = st.empty()
            insight_box = st.empty()
            
            current_insight = random.choice(INSIGHTS)
            insight_box.info(current_insight)
            
            status_text.markdown("### 🔍 Step 1/5: Analyzing question...")
            progress_bar.progress(5)
            time.sleep(0.3)
            
            status_text.markdown("### 🧠 Step 2/5: Detecting complexity...")
            progress_bar.progress(15)
            time.sleep(0.3)
            
            status_text.markdown("### 🤖 Step 3/5: Selecting AI model...")
            progress_bar.progress(25)
            time.sleep(0.3)
            
            status_text.markdown("### ⚡ Step 4/5: Generating SQL query...")
            progress_bar.progress(35)
            
            last_insight_time = time.time()
            insight_interval = 5
            query_start = time.time()
            query_done = False
            
            while not query_done:
                if time.time() - last_insight_time > insight_interval:
                    current_insight = random.choice([i for i in INSIGHTS if i != current_insight])
                    insight_box.info(current_insight)
                    last_insight_time = time.time()
                
                elapsed = time.time() - query_start
                simulated = min(70, 35 + int(elapsed * 0.5))
                progress_bar.progress(simulated)
                
                if result is None:
                    result = agent.ask(question)
                    query_done = True
                
                time.sleep(0.3)
            
            status_text.markdown("### ✅ Step 5/5: Formatting results...")
            progress_bar.progress(100)
            time.sleep(0.2)
        
        total_time = time.time() - start_time
        progress_container.empty()
        
        # ═══════════════════════════════════════════════════
        #  DISPLAY RESULTS
        # ═══════════════════════════════════════════════════
        
        if result["success"]:
            add_to_history(question, result, total_time)

            # Store last result for persistence
            st.session_state.last_result = result
            st.session_state.last_question = question
            st.session_state.last_time = total_time

            
            time_display = format_time(total_time)
            st.success(f"✅ Query completed in **{time_display}**")
            
            if total_time > 20 and result.get("models_tried"):
                with st.expander("⏱️ Why did it take this long?", expanded=True):
                    st.markdown("**📋 Model Execution Journey:**")
                    for i, m in enumerate(result["models_tried"], 1):
                        status = m["status"]
                        desc = m.get("description", m["model"])
                        err = m.get("error", "")
                        model_time = m.get("time", 0)
                        if "SUCCESS" in status:
                            st.markdown(f"**Step {i}:** {status} **{desc}** — ⏱️ {model_time}s")
                        elif "SKIPPED" in status:
                            st.markdown(f"**Step {i}:** {status} **{desc}** — ↳ {err}")
                        else:
                            st.markdown(f"**Step {i}:** {status} **{desc}** — ↳ {err[:100]}... ({model_time}s)")
                    st.markdown(f"---\n**Total:** {len(result['models_tried'])} models | **Used:** {result.get('model_used', 'Unknown')} | **Time:** {time_display}")
            
            st.markdown("---")
            st.markdown("### 💬 Answer")
            st.markdown(result["answer"])
            
            if total_time <= 20:
                st.caption(f"🤖 Model: {result.get('model_used', 'Unknown')}")
            
            with st.expander("🔍 View SQL Query"):
                st.code(result["query"], language="sql")
            
            if result.get("explanation"):
                with st.expander("📖 How This Query Works"):
                    st.markdown(result["explanation"])
            
            if result["results"]:
                with st.expander(f"📊 View Data ({result['row_count']} rows)", expanded=False):
                    df = pd.DataFrame(result["results"])
                    st.dataframe(df, use_container_width=True)
                    
                    # ─────────────────────────────────────
                    # EXPORT OPTIONS
                    # ─────────────────────────────────────
                    st.markdown("**📥 Export Options:**")
                    
                    export_col1, export_col2, export_col3, export_col4 = st.columns(4)
                    
                    # CSV Export
                    with export_col1:
                        csv_data = df.to_csv(index=False)
                        st.download_button(
                            "📄 CSV",
                            data=csv_data,
                            file_name=f"nexusiq_query_{int(time.time())}.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                    
                    # JSON Export
                    with export_col2:
                        json_data = df.to_json(orient="records", indent=2)
                        st.download_button(
                            "📋 JSON",
                            data=json_data,
                            file_name=f"nexusiq_query_{int(time.time())}.json",
                            mime="application/json",
                            use_container_width=True
                        )
                    
                    # Excel Export
                    with export_col3:
                        import io
                        from openpyxl.utils import get_column_letter
    
                        excel_buffer = io.BytesIO()
                        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                            df.to_excel(writer, index=False, sheet_name='Query Results')
                            
                            # Auto-adjust column widths
                            worksheet = writer.sheets['Query Results']
                            for col_idx, column in enumerate(df.columns, 1):
                                # Calculate max width needed
                                max_length = len(str(column))  # Header width
                                for value in df[column]:
                                    max_length = max(max_length, len(str(value)))
                                
                                # Set width with padding
                                adjusted_width = min(max_length + 3, 50)  # Cap at 50
                                worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                        excel_data = excel_buffer.getvalue()
                        st.download_button(
                            "📊 Excel",
                            data=excel_data,
                            file_name=f"nexusiq_query_{int(time.time())}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
                    
                    # Copy to Clipboard (Markdown Table)
                    with export_col4:
                        markdown_table = df.to_markdown(index=False)
                        st.download_button(
                            "📝 Markdown",
                            data=markdown_table,
                            file_name=f"nexusiq_query_{int(time.time())}.md",
                            mime="text/markdown",
                            use_container_width=True
                        )
        
        else:
            add_to_history(question, result, total_time)
            
            # VALIDATION ERRORS
            if result.get("validation_issues"):
                st.warning("⚠️ Question needs clarification")
                
                for issue in result["validation_issues"]:
                    issue_type = issue["type"]
                    details = issue["details"]
                    
                    if issue_type == "typo":
                        corrected = question.replace(details['typo'], details['suggestion'])
                        st.markdown(f"""
**Possible typo in {issue['field']}:**
- You wrote: `{details['typo']}`
- Did you mean: `{details['suggestion']}`?

**Available {issue['field']}s:** {', '.join(details['available'])}
                        """)
                        
                        st.success(f"✅ Try this: **{corrected}**")
                    
                    elif issue_type == "date_range":
                        st.markdown(f"""
**Date range issue:**
- {details['issue']}
- Available data: {details['data_range']}
- {details['suggestion']}
                        """)
                    
                    elif issue_type == "ambiguous":
                        st.markdown(f"**Ambiguous question:** {details['question']}")
                        
                        for i, option in enumerate(details['options']):
                            metric = option.split("(")[0].strip().lower().replace("by ", "")
                            st.success(f"✅ Try: **{question} by {metric}**")
                
                if result.get("suggestions"):
                    st.info("💡 **Suggestions:**\n\n" + "\n".join([f"• {s}" for s in result["suggestions"]]))
            
            else:
                st.error(f"❌ Query failed after {format_time(total_time)}")
                st.markdown(f"**Error:** {result.get('error', 'Unknown error')}")
            
            if result.get("models_tried"):
                with st.expander("📋 Model Journey"):
                    for m in result["models_tried"]:
                        st.markdown(f"• {m['status']} {m['model']}")
            
            st.info("💡 Try rephrasing or simplifying your question")
    
    elif ask_button and not question:
        st.warning("⚠️ Please enter a question")
        
    # ═══════════════════════════════════════════════════════
    #  SHOW LAST RESULT (persists after download clicks)
    # ═══════════════════════════════════════════════════════
    
    elif not ask_button and st.session_state.last_result and not st.session_state.selected_history:
        result = st.session_state.last_result
        question = st.session_state.last_question
        total_time = st.session_state.last_time
        
        time_display = format_time(total_time)
        st.success(f"✅ Query completed in **{time_display}**")
        
        if total_time > 20 and result.get("models_tried"):
            with st.expander("⏱️ Why did it take this long?", expanded=True):
                for i, m in enumerate(result["models_tried"], 1):
                    st.markdown(f"**Step {i}:** {m['status']} {m.get('description', m['model'])} ({m['time']}s)")
        
        st.markdown("---")
        st.markdown("### 💬 Answer")
        st.markdown(result["answer"])
        
        if total_time <= 20:
            st.caption(f"🤖 Model: {result.get('model_used', 'Unknown')}")
        
        with st.expander("🔍 View SQL Query"):
            st.code(result["query"], language="sql")
        
        if result.get("explanation"):
            with st.expander("📖 How This Query Works"):
                st.markdown(result["explanation"])
        
        if result["results"]:
            with st.expander(f"📊 View Data ({result['row_count']} rows)", expanded=False):
                df = pd.DataFrame(result["results"])
                st.dataframe(df, use_container_width=True)
                
                st.markdown("**📥 Export Options:**")
                export_col1, export_col2, export_col3, export_col4 = st.columns(4)
                
                with export_col1:
                    csv_data = df.to_csv(index=False)
                    st.download_button(
                        "📄 CSV",
                        data=csv_data,
                        file_name=f"nexusiq_query_{int(time.time())}.csv",
                        mime="text/csv",
                        use_container_width=True,
                        key="persist_csv"
                    )
                
                with export_col2:
                    json_data = df.to_json(orient="records", indent=2)
                    st.download_button(
                        "📋 JSON",
                        data=json_data,
                        file_name=f"nexusiq_query_{int(time.time())}.json",
                        mime="application/json",
                        use_container_width=True,
                        key="persist_json"
                    )
                
                with export_col3:
                    import io
                    from openpyxl.utils import get_column_letter
                    
                    excel_buffer = io.BytesIO()
                    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Query Results')
                        worksheet = writer.sheets['Query Results']
                        for col_idx, column in enumerate(df.columns, 1):
                            max_length = len(str(column))
                            for value in df[column]:
                                max_length = max(max_length, len(str(value)))
                            adjusted_width = min(max_length + 3, 50)
                            worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                    
                    st.download_button(
                        "📊 Excel",
                        data=excel_buffer.getvalue(),
                        file_name=f"nexusiq_query_{int(time.time())}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key="persist_excel"
                    )
                
                with export_col4:
                    markdown_table = df.to_markdown(index=False)
                    st.download_button(
                        "📝 Markdown",
                        data=markdown_table,
                        file_name=f"nexusiq_query_{int(time.time())}.md",
                        mime="text/markdown",
                        use_container_width=True,
                        key="persist_md"
                    )
    
    st.markdown("---")
    st.caption("🧠 NexusIQ AI | Intelligent multi-model fallback with quota tracking")